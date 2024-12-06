from django.shortcuts import render, get_object_or_404, redirect
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook import Workbook

from .models import Salary
from django.urls import reverse
from django.core.paginator import Paginator
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
# from .forms import SalaryForm
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta, time
from attendance.models import AttendanceTracking
from staff.models import StaffProfile, Department
from authentication.decorators import admin_required
from django.db.models import Q



# Quản lý lương
@admin_required
def salary_list(request):
    # Xử lý tìm kiếm và lọc
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    per_page = request.GET.get('per_page', 10)

    current_month = timezone.now().month
    month_filter = request.GET.get('month', current_month)


    salaries = Salary.objects.select_related(
        'staff',
        'staff__department',
        'staff__position',
        'staff__user'
    )
    
    # Áp dụng bộ lọc tìm kiếm
    if search_query:
        salaries = salaries.filter(
            Q(staff__id__icontains=search_query) |  # Tìm theo mã nhân viên
            Q(staff__full_name__icontains=search_query) |  # Tìm theo tên
            Q(staff__user__email__icontains=search_query) |  # Tìm theo email
            Q(staff__phone_number__icontains=search_query) |  # Tìm theo số điện thoại
            Q(staff__department__name__icontains=search_query) |  # Tìm theo tên phòng ban
            Q(staff__position__name__icontains=search_query)  # Tìm theo tên chức vụ
        )
    
    # Lọc theo phòng ban
    if department_filter:
        salaries = salaries.filter(staff__department_id=department_filter)
    
    # Lọc theo tháng
    if month_filter:
        salaries = salaries.filter(pay_date__month=month_filter)

    # Phân trang
    paginator = Paginator(salaries, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Lấy danh sách phòng ban cho dropdown
    departments = Department.objects.all()
    
    today = datetime.now()
    months = range(1, 13)

    context = {
        'page_obj': page_obj,
        'departments': departments,
        'search_query': search_query,
        'department_filter': department_filter,
        'months': months,
        'month_filter': int(month_filter) if month_filter else current_month,
        'year_filter': today.year,
        'current_month': today.month,
        'current_year': today.year,
        'per_page': int(per_page)
    }
    
    return render(request, 'salary_list.html', context)

@login_required
def export_salary(request):
    try:
        # Lấy các tham số filter
        today = datetime.now()
        search_query = request.GET.get('search', '')
        department_filter = request.GET.get('department', '')
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        # Xử lý tháng/năm mặc định
        try:
            month = int(month) if month else today.month
            year = int(year) if year else today.year
        except ValueError:
            month = today.month
            year = today.year
            
        # Validate month/year
        if not (1 <= month <= 12) or year < 2000:
            month = today.month
            year = today.year
        
        # Query cơ sở với select_related
        salaries = Salary.objects.select_related(
            'staff',
            'staff__department',
            'staff__position'
        ).filter(
            pay_date__month=month,
            pay_date__year=year
        )
        
        # Áp dụng các bộ lọc tìm kiếm
        if search_query:
            salaries = salaries.filter(
                Q(staff__id__icontains=search_query) |
                Q(staff__full_name__icontains=search_query) |
                Q(staff__user__email__icontains=search_query) |
                Q(staff__phone_number__icontains=search_query) |
                Q(staff__department__name__icontains=search_query) |
                Q(staff__position__name__icontains=search_query)
            )
        
        # Lọc theo phòng ban
        if department_filter:
            salaries = salaries.filter(staff__department_id=department_filter)
        
        # Tạo workbook và sheet
        wb = Workbook()
        sheet_name = f"Luong_{month}_{year}"
        ws = wb.active
        ws.title = sheet_name
        
        # Merge cells cho tiêu đề
        ws.merge_cells('A1:I1')  # Merge từ cột A đến I cho tiêu đề chính
        ws.merge_cells('A2:I2')  # Merge cho thông tin tháng/năm
        
        # Thêm tiêu đề chính
        ws['A1'] = 'BÁO CÁO LƯƠNG NHÂN VIÊN'
        ws['A2'] = f'Tháng {month}/{year}'
        
        current_row = 3  # Bắt đầu từ dòng 3
        
        # Thông tin filter (nếu có)
        if department_filter or search_query:
            if department_filter:
                ws.merge_cells(f'A{current_row}:I{current_row}')
                department = Department.objects.get(id=department_filter)
                ws[f'A{current_row}'] = f'Phòng ban: {department.name}'
                current_row += 1
            if search_query:
                ws.merge_cells(f'A{current_row}:I{current_row}')
                ws[f'A{current_row}'] = f'Tìm kiếm: {search_query}'
                current_row += 1
            current_row += 1  # Thêm dòng trống
        else:
            current_row += 1  # Thêm dòng trống nếu không có filter
        
        # Style cho phần header
        header_font = Font(bold=True, size=14)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Áp dụng style cho tiêu đề và thông tin filter
        for row in range(1, current_row):
            cell = ws[f'A{row}']
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Tạo header cho bảng dữ liệu
        headers = [
            'Mã NV', 'Họ Tên', 'Phòng Ban', 'Chức Vụ',
            'Lương Cơ Bản', 'Số Ngày Công', 'Thưởng',
            'Ngày Trả Lương', 'Tổng Lương'
        ]
        
        # Thêm header bảng
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Thêm dữ liệu
        for salary in salaries:
            ws.append([
                salary.staff.id or '',
                salary.staff.full_name or '',
                salary.staff.department.name if salary.staff.department else '',
                salary.staff.position.name if salary.staff.position else '',
                salary.staff.base_salary or 0,
                salary.num_days_work or 0,
                salary.bonus or 0,
                salary.pay_date.strftime('%d/%m/%Y') if salary.pay_date else '',
                salary.total_salary or 0
            ])
        
        # Style cho dữ liệu
        data_alignment = Alignment(horizontal='center')
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=current_row + 1):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = data_border
        
        # Format số tiền (bắt đầu từ dòng sau header)
        money_format = '#,##0'
        for row in ws.iter_rows(min_row=current_row + 1):
            row[4].number_format = money_format  # Lương cơ bản
            row[6].number_format = money_format  # Thưởng
            row[8].number_format = money_format  # Tổng lương
            
        # Điều chỉnh độ rộng cột
        fixed_widths = {
            'A': 10,  # Mã NV
            'B': 25,  # Họ Tên
            'C': 20,  # Phòng Ban
            'D': 20,  # Chức Vụ
            'E': 15,  # Lương Cơ Bản
            'F': 12,  # Số Ngày Công
            'G': 15,  # Thưởng
            'H': 15,  # Ngày Trả Lương
            'I': 15,  # Tổng Lương
        }
        for col_letter, width in fixed_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Tạo response
        filename = f"Bang_luong_{month}_{year}"
        if department_filter:
            department = Department.objects.get(id=department_filter)
            filename += f"_{department.name}"
        if search_query:
            filename += "_filtered"
        filename += ".xlsx"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def generate_salary(request):
    try:
        # Lấy tháng hiện tại
        today = datetime.now()
        month = today.month
        year = today.year
        
        # Kiểm tra xem đã tạo lương tháng này chưa
        if Salary.objects.filter(
            pay_date__month=month,
            pay_date__year=year
        ).exists():
            return JsonResponse({
                'status': 'error',
                'message': f'Bảng lương tháng {month}/{year} đã được tạo!'
            }, status=400)

        # Lấy danh sách nhân viên
        staffs = StaffProfile.objects.all()
        salaries_created = 0
        
        for staff in staffs:
            # Lấy tất cả bản ghi chấm công trong tháng
            attendance_records = AttendanceTracking.objects.filter(
                staff=staff,
                date_work__month=month,
                date_work__year=year,
                checkout__isnull=False  # Chỉ tính những ngày đã checkout
            )

            # Tính toán các chỉ số
            work_stats = calculate_work_statistics(attendance_records)
            
            # Tính thưởng
            bonus = calculate_bonus(
                work_stats['normal_days'],
                work_stats['overtime_hours'],
                work_stats['late_days'],
                staff.base_salary
            )
            
            # Tạo bản ghi lương mới
            Salary.objects.create(
                staff=staff,
                num_days_work=work_stats['normal_days'],
                bonus=bonus,
                pay_date=today,
                total_salary=calculate_total_salary(
                    staff.base_salary,
                    work_stats,
                    bonus
                )
            )
            salaries_created += 1

        return JsonResponse({
            'status': 'success',
            'message': f'Đã tạo {salaries_created} bảng lương tháng {month}/{year} thành công!',
            'data': {
                'month': month,
                'year': year,
                'salaries_created': salaries_created,
                'created_at': today.strftime('%d/%m/%Y %H:%M:%S')
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

def calculate_work_statistics(attendance_records):
    """
    Tính toán thống kê làm việc:
    - Số ngày làm việc bình thường
    - Số giờ tăng ca
    - Số ngày đi muộn
    """
    normal_days = 0
    overtime_hours = 0
    late_days = 0
    
    STANDARD_HOURS = 8  # Số giờ làm việc tiêu chuẩn
    LATE_THRESHOLD = timedelta(minutes=15)  # Ngưỡng đi muộn (15 phút)

    for record in attendance_records:
        # Tính thời gian làm việc trong ngày
        if record.checkout and record.checkin:
            
            # Kiểm tra đi muộn
            if record.checkin > (datetime.combine(datetime.min, time(8, 0)) + LATE_THRESHOLD).time():
                late_days += 1
            
            # Tính ngày làm việc và giờ tăng ca
            if record.total_working >= STANDARD_HOURS:
                normal_days += 1
                if record.overtime > 0:
                    overtime_hours += record.overtime
    
    return {
        'normal_days': normal_days,
        'overtime_hours': overtime_hours,
        'late_days': late_days
    }

def calculate_bonus(normal_days, overtime_hours, late_days, base_salary):
    """
    Tính thưởng dựa trên:
    - Số ngày làm việc bình thường
    - Số giờ tăng ca
    - Số ngày đi muộn
    - Lương cơ bản
    """
    bonus = 0
    standard_days = 22
    
    # Thưởng đi làm đủ ngày
    if normal_days >= standard_days:
        bonus += base_salary * 0.1  # Thưởng 10% lương cơ bản
    
    # Thưởng tăng ca
    overtime_rate = base_salary / (standard_days * 8) * 1.5  # Lương tăng ca = 150% lương giờ bình thường
    bonus += overtime_hours * overtime_rate
    
    # Trừ phạt đi muộn
    late_penalty = base_salary / standard_days * 0.1  # Phạt 10% lương ngày
    bonus -= late_days * late_penalty
    
    return round(max(bonus, 0))  # Đảm bảo thưởng không âm

def calculate_total_salary(base_salary, work_stats, bonus):
    """
    Tính tổng lương dựa trên:
    - Lương cơ bản
    - Thống kê làm việc
    - Thưởng
    """
    standard_days = 22
    daily_salary = base_salary / standard_days
    
    # Lương cơ bản theo ngày làm việc
    working_salary = daily_salary * work_stats['normal_days']
    
    # Tổng lương = Lương làm việc + Thưởng
    total = working_salary + bonus
    
    return round(total)

@login_required
def staff_salary_list(request):
    # Lấy nhân viên hiện tại
    staff = request.user.staff_profile
    per_page = request.GET.get('per_page', 10)

    # Lấy tháng hiện tại
    current_month = timezone.now().month
    months = range(1, 13)
    month_filter = request.GET.get('month', current_month)

    # Query lương của nhân viên
    salaries = Salary.objects.filter(staff=staff).order_by('-pay_date')

    # Lọc theo tháng nếu có
    if month_filter:
        salaries = salaries.filter(pay_date__month=month_filter)

    # Phân trang
    paginator = Paginator(salaries, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'months': months,
        'month_filter': int(month_filter) if month_filter else current_month,
        'per_page': int(per_page)
    }
    return render(request, 'staff_salary_list.html', context)